"""
FastAPI application for XLSX pivot service.
Handles file uploads, processes student scores, and returns consolidated reports.
"""

import io
import logging
from typing import Dict, Any
from datetime import datetime

from fastapi import FastAPI, File, UploadFile, HTTPException, status
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from .parser import process_student_workbook

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Initialize FastAPI app
app = FastAPI(
    title="XLSX Pivot Service for Student Scores",
    description="Automated processing of Excel files containing student scores and discipline lists",
    version="1.0.0"
)

# Configure CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Configure appropriately for production
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def highlight_problematic_cells(output_bytes: bytes) -> bytes:
    """
    Highlights cells with missing values, '!', or '?' symbols in the Excel file.
    
    Args:
        output_bytes: Bytes content of the Excel file
        
    Returns:
        Modified bytes with highlighted cells
    """
    # Load the workbook
    wb = load_workbook(io.BytesIO(output_bytes))
    
    # Define fill colors
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow for missing
    red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')     # Light red for ! and ?
    
    # Process Report sheet
    if 'Report' in wb.sheetnames:
        ws = wb['Report']
        logger.info(f"Processing Report sheet with {ws.max_row} rows and {ws.max_column} columns")
        
        # Start from row 2 (skip header) and column 2 (skip index)
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(2, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_value = cell.value
                
                # Check for problematic values
                if cell_value is None or cell_value == '' or (isinstance(cell_value, str) and cell_value.strip() == ''):
                    # Missing value - highlight in yellow
                    cell.fill = yellow_fill
                    logger.debug(f"Highlighted empty cell at row {row_idx}, col {col_idx}")
                    
                elif isinstance(cell_value, str) and ('!' in cell_value or '?' in cell_value):
                    # Contains ! or ? - highlight in red
                    cell.fill = red_fill
                    logger.debug(f"Highlighted cell with '{cell_value}' at row {row_idx}, col {col_idx}")
                elif isinstance(cell_value, str) and cell_value.strip() != '':
                    # Convert string digits to int if possible
                    try:
                        cell.value = int(cell_value)
                        logger.debug(f"Converted cell value '{cell_value}' to integer at row {row_idx}, col {col_idx}") 
                    except ValueError:
                        pass  # Not a digit, skip
    
    # Process RawScores sheet
    if 'RawScores' in wb.sheetnames:
        ws = wb['RawScores']
        logger.info(f"Processing RawScores sheet with {ws.max_row} rows and {ws.max_column} columns")
        
        # Start from row 2 (skip header) and column 2 (skip index)
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(2, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_value = cell.value
                
                # Check for problematic values
                if cell_value is None or cell_value == '' or (isinstance(cell_value, str) and cell_value.strip() == ''):
                    # Missing value - highlight in yellow
                    cell.fill = yellow_fill
                    
                elif isinstance(cell_value, str) and ('!' in cell_value or '?' in cell_value):
                    # Contains ! or ? - highlight in red
                    cell.fill = red_fill
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    logger.info("Cell highlighting completed")
    return output.getvalue()


@app.get("/")
async def root() -> Dict[str, str]:
    """Root endpoint returning service information."""
    return {
        "service": "XLSX Pivot Service",
        "version": "1.0.0",
        "status": "operational"
    }


@app.get("/health")
async def health() -> Dict[str, str]:
    """Health check endpoint."""
    return {
        "status": "healthy",
        "timestamp": datetime.utcnow().isoformat()
    }


@app.post("/pivot")
async def create_pivot(
    scores_xlsx: UploadFile = File(..., description="Student scores workbook (XLSX)"),
    disciplines_xlsx: UploadFile = File(..., description="Discipline list workbook (XLSX)")
) -> StreamingResponse:
    """
    Process student scores and create a consolidated pivot report.
    
    Args:
        scores_xlsx: Excel file with student sheets and grades
        disciplines_xlsx: Excel file with official discipline list
        
    Returns:
        StreamingResponse: Generated XLSX file with Report and RawScores sheets
        
    Raises:
        HTTPException: If file validation or processing fails
    """
    logger.info("Received pivot request")
    
    # Validate file types
    if not scores_xlsx.filename.endswith('.xlsx'):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="scores_xlsx must be an Excel file (.xlsx)"
        )
    
    if not disciplines_xlsx.filename.endswith('.xlsx'):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="disciplines_xlsx must be an Excel file (.xlsx)"
        )
    
    try:
        # Read uploaded files into memory
        logger.info(f"Reading scores file: {scores_xlsx.filename}")
        scores_content = await scores_xlsx.read()
        
        logger.info(f"Reading disciplines file: {disciplines_xlsx.filename}")
        disciplines_content = await disciplines_xlsx.read()
        
        # Validate file sizes (basic check)
        if len(scores_content) == 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="scores_xlsx file is empty"
            )
        
        if len(disciplines_content) == 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="disciplines_xlsx file is empty"
            )
        
        logger.info("Processing student workbook")
        logger.info(f"Len of scores_content {len(scores_content)}")
        logger.info(f"Len of disciplines_content {len(disciplines_content)}")
        
        # Process the workbooks
        df_raw_scores, df_report = process_student_workbook(
            scores_content,
            disciplines_content
        )
        
        logger.info("Generating output XLSX")
        
        # Create output Excel file in memory
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write Report sheet (final matched disciplines)
            df_report.to_excel(writer, sheet_name='Report', index=True)
            
            # Write RawScores sheet (intermediate normalized data)
            df_raw_scores.to_excel(writer, sheet_name='RawScores', index=True)
        
        output.seek(0)
        
        # Apply highlighting to problematic cells
        logger.info("Applying cell highlighting")
        highlighted_bytes = highlight_problematic_cells(output.getvalue())
        
        # Create new BytesIO with highlighted content
        final_output = io.BytesIO(highlighted_bytes)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"pivot_report_{timestamp}.xlsx"
        
        logger.info(f"Returning result: {filename}")
        
        return StreamingResponse(
            final_output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
    
    except HTTPException:
        # Re-raise HTTP exceptions
        raise
    
    except pd.errors.EmptyDataError as e:
        logger.error(f"Empty data error: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="One or more Excel files contain no data"
        )
    
    except ValueError as e:
        logger.error(f"Validation error: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Data validation failed: {str(e)}"
        )
    
    except Exception as e:
        logger.error(f"Processing error: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to process workbooks: {str(e)}"
        )


@app.post("/validate")
async def validate_files(
    scores_xlsx: UploadFile = File(..., description="Student scores workbook (XLSX)"),
    disciplines_xlsx: UploadFile = File(..., description="Discipline list workbook (XLSX)")
) -> Dict[str, Any]:
    """
    Validate uploaded Excel files without processing.
    
    Args:
        scores_xlsx: Excel file with student sheets
        disciplines_xlsx: Excel file with discipline list
        
    Returns:
        Validation results including sheet counts and basic structure info
    """
    logger.info("Received validation request")
    
    try:
        scores_content = await scores_xlsx.read()
        disciplines_content = await disciplines_xlsx.read()
        
        # Read scores workbook
        scores_excel = pd.ExcelFile(io.BytesIO(scores_content), engine='openpyxl')
        scores_sheets = scores_excel.sheet_names
        
        # Read disciplines workbook
        disciplines_excel = pd.ExcelFile(io.BytesIO(disciplines_content), engine='openpyxl')
        disciplines_sheets = disciplines_excel.sheet_names
        
        # Basic validation checks
        validation_result = {
            "valid": True,
            "scores_file": {
                "filename": scores_xlsx.filename,
                "sheet_count": len(scores_sheets),
                "sheet_names": scores_sheets,
                "size_bytes": len(scores_content)
            },
            "disciplines_file": {
                "filename": disciplines_xlsx.filename,
                "sheet_count": len(disciplines_sheets),
                "sheet_names": disciplines_sheets,
                "size_bytes": len(disciplines_content)
            },
            "warnings": []
        }
        
        # Add warnings if needed
        if len(scores_sheets) == 0:
            validation_result["warnings"].append("Scores workbook has no sheets")
            validation_result["valid"] = False
        
        if len(disciplines_sheets) == 0:
            validation_result["warnings"].append("Disciplines workbook has no sheets")
            validation_result["valid"] = False
        
        logger.info(f"Validation complete: {validation_result['valid']}")
        return validation_result
    
    except Exception as e:
        logger.error(f"Validation error: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"File validation failed: {str(e)}"
        )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)