"""
Standalone script to highlight problematic cells in an Excel file.
Highlights:
- Missing values (empty cells) in YELLOW
- Cells containing '!' or '?' in RED
"""

import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def highlight_problematic_cells(input_file: str, output_file: str):
    """
    Highlights cells with missing values, '!', or '?' symbols in the Excel file.
    
    Args:
        input_file: Path to input Excel file
        output_file: Path to output Excel file with highlights
    """
    print(f"Loading workbook: {input_file}")
    wb = load_workbook(input_file)
    
    # Define fill colors
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow for missing
    red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')     # Light red for ! and ?
    
    stats = {
        'missing': 0,
        'symbols': 0,
        'total_cells': 0
    }
    
    # Process all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nProcessing sheet: {sheet_name}")
        print(f"  Dimensions: {ws.max_row} rows x {ws.max_column} columns")
        
        # Start from row 2 (skip header) and column 2 (skip index if present)
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(2, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_value = cell.value
                stats['total_cells'] += 1
                
                # Check for problematic values
                if cell_value is None or cell_value == '' or (isinstance(cell_value, str) and cell_value.strip() == ''):
                    # Missing value - highlight in yellow
                    cell.fill = yellow_fill
                    stats['missing'] += 1
                    
                elif isinstance(cell_value, str) and ('!' in cell_value or '?' in cell_value):
                    # Contains ! or ? - highlight in red
                    cell.fill = red_fill
                    stats['symbols'] += 1
                    print(f"  Found problematic value '{cell_value}' at row {row_idx}, col {col_idx}")
    
    # Save the workbook
    print(f"\nSaving highlighted workbook to: {output_file}")
    wb.save(output_file)
    
    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    print(f"Total cells checked: {stats['total_cells']}")
    print(f"Missing values (yellow): {stats['missing']}")
    print(f"Cells with ! or ? (red): {stats['symbols']}")
    print(f"\nOutput saved to: {output_file}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python highlight_cells.py <input_file.xlsx> <output_file.xlsx>")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2]
    
    highlight_problematic_cells(input_file, output_file)